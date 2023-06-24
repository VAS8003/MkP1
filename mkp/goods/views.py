import os
from django.contrib.auth import logout, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from django.contrib import messages
from django.http import HttpResponse, FileResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.utils.text import slugify
from django.views.generic import ListView, DeleteView, CreateView
from django.contrib.auth.mixins import LoginRequiredMixin
import pandas as pd
from wsgiref.util import FileWrapper
from opt.utils import DataSellerMixin
from .forms import *
from .utils import *
import openpyxl
from django.views import View
from .models import Good, Category
from django.shortcuts import render
from django.core.files.storage import default_storage


def search_goods(request):
    search_query = request.GET.get('search')
    if search_query:
        goods = Good.objects.filter(title__icontains=search_query)
    else:
        goods = Good.objects.all()

    context = {'goods': goods}
    return render(request, 'update_export_file.html', context)


@user_passes_test(lambda u: u.username == 'adminVAS80', login_url='login')
def update_export_file(request):
    file_path = 'export.xlsx'
    if default_storage.exists(file_path):
        # Видаляємо існуючий файл
        default_storage.delete(file_path)

    # Завантажуємо новий файл
    file = request.FILES.get('file')
    if file:
        default_storage.save(file_path, file)

    context = {'file_path': file_path}
    return render(request, 'goods/update_export_file.html', context)


class ImportView(LoginRequiredMixin, View):
    template_name = 'goods/import.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        if 'file' not in request.FILES:
            return render(request, self.template_name, {'error': 'No file selected'})

        file = request.FILES['file']
        if not file.name.endswith('.xlsx'):
            return render(request, self.template_name, {'error': 'Invalid file format'})

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            errors = []
            success_count = 0

            for row in ws.iter_rows(values_only=True):
                article = row[10]  # Артикул товару

                try:
                    existing_good = Good.objects.get(article=article)
                except Good.DoesNotExist:
                    # Якщо товар не знайдено, створюємо новий товар
                    try:
                        cat_name = row[18]  # Назва категорії товару
                        brand_name = row[19]  # Назва бренду товару

                        # Перевіряємо наявність категорії та бренду
                        category, _ = Category.objects.get_or_create(name=cat_name)
                        brand, _ = Brand.objects.get_or_create(name=brand_name)
                        title = row[0]  # Назва товару
                        slug = slugify(title)  # Генеруємо slug на основі URL
                        photo_data = row[20]  # Значення фотографії у файлі

                        new_good = Good(
                            title=row[0],  # Найменування
                            slug=slug,  # URL
                            annotation=row[2],  # Короткий опис
                            description=row[3],  # Опис
                            article=article,  # Артикул
                            price_opt=row[11],  # Ціна ОПТ
                            price_b2c=row[12],  # Ціна Роздріб
                            price_drop=row[13],  # Ціна DROP
                            price_mrc=row[14],  # МРЦ
                            cat=category,  # Категорія
                            brand=brand,
                            owner=request.user,
                            provider=row[8],
                            full_stock=row[9],
                            opt_stock=row[15],
                            b2c_stock=row[16],
                            drop_stock=row[16],
                            photo=photo_data,
                        )

                        # Перевірка на наявність помилок у новому товарі перед збереженням
                        try:
                            new_good.full_clean()
                            new_good.save()
                            success_count += 1
                        except ValidationError as e:
                            errors.append(f"Помилка при створенні товару: {e}")

                    except Exception as e:
                        # Обробка помилки при створенні товару
                        errors.append(f"Помилка при створенні товару: {e}")

            success_message = f"Успішно імпортовано {success_count} товарів"
            context = {'success': success_message, 'errors': errors}
            return render(request, self.template_name, context)
        except Exception as e:
            # Обробка помилки при завантаженні файлу
            return render(request, self.template_name, {'error': f'Error loading file: {e}'})



class ExportView(View):
    def get(self, request):
        # Отримуємо всі об'єкти моделі Good
        goods = Good.objects.all()
        goods = Good.objects.all()

        # Створюємо новий документ Excel
        wb = openpyxl.Workbook()
        ws = wb.active

        # Додаємо заголовки стовпців
        headers = ['ID', 'Найменування', 'URL', 'Короткий опис', 'Опис', 'Створено', 'Оновлено', 'Опубліковано',
                   'Створив', 'Поставщик', 'Стандарт наявності', 'Артикул', 'Ціна ОПТ', 'Ціна Роздріб', 'Ціна DROP',
                   'МРЦ', 'Опт наявність', 'Роздріб наявність', 'DROP наявність', 'Категорія', 'Бренд']
        ws.append(headers)

        # Додаємо дані для кожного товару
        for good in goods:
            row = [
                good.id,
                good.title,
                good.slug,
                good.annotation,
                good.description,
                good.time_create.replace(tzinfo=None),
                good.time_update.replace(tzinfo=None),
                good.is_published,
                good.owner.username,
                good.provider,
                good.full_stock,
                good.article,
                good.price_opt,
                good.price_b2c,
                good.price_drop,
                good.price_mrc,
                good.opt_stock,
                good.b2c_stock,
                good.drop_stock,
                good.cat.name,
                good.brand.name if good.brand else ''
            ]
            ws.append(row)

        # Встановлюємо заголовки стовпців на першому рядку
        for col_num, column_title in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            cell = ws[f'{col_letter}1']
            cell.value = column_title
            cell.font = openpyxl.styles.Font(bold=True)

        # Налаштування HTTP-відповіді для файлу Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=goods.xlsx'

        # Збереження документа Excel у відповідь
        wb.save(response)

        return response




class ExportFileView(View):
    def get(self, request):
        file_path = r'C:\Users\Pavilion\PycharmProjects\MkP\mkp\media\export.xlsx'

        # Зчитуємо файл export.xlsx
        df = pd.read_excel(file_path)

        # Оновлюємо значення поля "Количество" на основі моделі Good
        for index, row in df.iterrows():
            article = row['Код_товара']
            try:
                good = Good.objects.get(article=article)
                quantity = int(good.opt_stock)
                if quantity == 0:
                    df.at[index, 'Наличие'] = "-"
                else:
                    df.at[index, 'Наличие'] = "+"
                df.at[index, 'Количество'] = quantity
            except Good.DoesNotExist:
                pass

        # Зберігаємо оновлений файл
        df.to_excel(file_path, index=False)

        # Відправляємо файл користувачеві
        if os.path.exists(file_path):
            file_wrapper = FileWrapper(open(file_path, 'rb'))
            response = FileResponse(file_wrapper)
            response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
            return response
        else:
            return HttpResponse('File not found', status=404)

@user_passes_test(lambda u: u.username == 'adminVAS80', login_url='login')
def update_stock(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                df = pd.read_excel(excel_file)

                for _, row in df.iterrows():
                    article = row['Артикул']
                    opt_stock = row['Залишок']

                    if pd.notnull(opt_stock):  # Перевіряємо, чи не є значення в полі "Залишок" пустим
                        Good.objects.filter(article=article).update(opt_stock=opt_stock)

                messages.success(request, 'Залишки оновлено успішно.')
            except Exception as e:
                messages.error(request, f'Не вдалося оновити залишки. Помилка: {str(e)}')

            return redirect('update_stock')

    return render(request, 'goods/update_stock.html')


class GoodsList(DataMixin,LoginRequiredMixin, ListView):
    model = Good
    template_name = 'goods/show_goods.html'
    context_object_name = 'goods'
    allow_empty = True
    login_url = 'login'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Всі товари")
        return dict(list(context.items()) + list(c_def.items()))

    def get_queryset(self):
        return Good.objects.filter(is_published=True, opt_stock__gte=1).select_related('cat')



class GoodCategory(DataMixin, ListView):
    model = Good
    template_name = 'goods/show_goods.html'
    context_object_name = 'goods'


    def get_queryset(self):
        return Good.objects.filter(cat__slug=self.kwargs['cat_slug'], is_published=True, opt_stock__gte=1).select_related('cat')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c = Category.objects.get(slug=self.kwargs['cat_slug'])
        c_def = self.get_user_context(title='Категорія - ' + str(c.name),
                                      cat_selected=c.pk)
        return dict(list(context.items()) + list(c_def.items()))


class GoodBrand(DataMixin, ListView):
    model = Good
    template_name = 'goods/show_goods.html'
    context_object_name = 'goods'


    def get_queryset(self):
        return Good.objects.filter(brand__slug=self.kwargs['brand_slug'], is_published=True, opt_stock__gte=1).select_related('brand')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        b = Brand.objects.get(slug=self.kwargs['brand_slug'])
        b_def = self.get_user_context(title=str(b.name),
                                      brand_selected=b.pk)
        return dict(list(context.items()) + list(b_def.items()))


def category(requests):
    pass

class ShowGood(LoginRequiredMixin, DataMixin, DeleteView):
    model = Good
    template_name = 'goods/good.html'
    slug_url_kwarg = 'good_slug'
    context_object_name = "good"
    allow_empty = True
    login_url = 'login'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title= context['good'])
        return dict(list(context.items()) + list(c_def.items()))




class AddGood(LoginRequiredMixin, DataSellerMixin, CreateView):
    form_class = AddGoodForms
    template_name = 'goods/add_good.html'
    success_url = reverse_lazy('seller')
    login_url = reverse_lazy('goods')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Добавлення товару")
        return dict(list(context.items()) + list(c_def.items()))

    def form_valid(self, form):
        form.instance.owner = self.request.user
        instance = form.save(commit=False)
        instance.owner = self.request.user
        instance.save()
        return super().form_valid(form)







class RegisterUser(DataMixin, CreateView):
    form_class = RegisterUserForm
    template_name = 'goods/register.html'
    success_url = reverse_lazy('login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title="Реєстрація")
        return dict(list(context.items()) + list(c_def.items()))

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        return redirect('goods')


class LoginUser(DataMixin, LoginView):
     form_clas = LoginUserForm
     template_name = 'goods/login.html'

     def get_context_data(self, *, object_list=None, **kwargs):
         context = super().get_context_data(**kwargs)
         c_def = self.get_user_context(title="Авторизація")
         return dict(list(context.items()) + list(c_def.items()))

     def get_success_url(self):
         return reverse_lazy('goods')


def logout_user(requests):
    logout(requests)
    return redirect('login')

